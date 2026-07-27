#!/usr/bin/env python3
# coding: utf-8
"""
Генератор базы яхт для сервиса бронирования.

Схема колонок задана вручную в «Яхты бронирование.xlsx» — здесь она
воспроизводится один в один. Шесть строк из исходного файла переносятся
без изменений и служат эталоном: все выдуманные лодки считаются по
формулам, откалиброванным по ним.

Запуск:  python3 generate.py      (нужен openpyxl)
Выход:   yachts.xlsx, yachts.csv, PHOTO-CREDITS.md

Данные вымышленные — это учебный проект. Реальны только названия
регионов, портов и моделей лодок, а также ссылки на фотографии.
"""

import csv
import json
import os
import random

HERE = os.path.dirname(os.path.abspath(__file__))

COLUMNS = [
    "Название", "Тип Яхты", "Команда", "Добавить повара", "Регион", "Порт",
    "Каюты", "Длина", "Места для гостей", "Туалет ", "Стоимость за день",
    "Состояние", "Фото",
]

# --- Строки из исходной таблицы. Переносятся дословно. ----------------------
SEED_ROWS = [
    ["Белая ночь", "Полутонник", "Капитан и помощник", "Нет", "Белое море", "Чупа",
     2, "10 м", 5, "Нет", 15000, 3,
     "https://static.tildacdn.com/tild3163-6436-4461-b630-333533373061/Photo_21052021_19_55.jpg"],
    ["Морти", "Bavaria 45", "Капитан", "Да", "Владивосток", "Центральный",
     5, "15 м", 10, 2, 35000, 5,
     "https://static.tildacdn.com/tild6266-3762-4766-b939-303630376638/_DSC1609.jpg"],
    ["Liberty", "McGregor 65", "Капитан и помощник", "Да", "Камчатка", "Фриза",
     6, "22 м", 8, 1, 30000, 4,
     "https://static.tildacdn.com/tild3335-3634-4266-b932-383262303338/DSC09472.jpg"],
    ["Яхта", "Конрад 25RT", "Капитан", "Нет", "Ладога", "Сортавала",
     3, "9 м", 6, "Нет", 16000, 3, ""],
    ["Рик", "Bavaria 45", "Капитан", "Да", "Владивосток", "Центральный",
     5, "15 м", 10, 2, 35000, 5, ""],
    ["Iron Lady", "Судоход 115", "Капитан и помощник", "Да", "Камчатка", "Фриза",
     8, "28 м", 16, 2, 64000, 5, ""],
]

# Снимок до того, как строкам без фото раздадут кадры: списки SEED_ROWS
# правятся на месте, и после этого отличить исходные ссылки уже нельзя.
TILDA_PHOTOS = [r[12] for r in SEED_ROWS if r[12]]

# --- Сорок направлений: регион, порт, ценовой коэффициент -------------------
# Коэффициент отражает логистику: чем дальше и безлюднее берег, тем дороже
# подать туда лодку и экипаж. Азов и Каспий — самые доступные, Арктика и
# Командоры — самые дорогие.
REGIONS = [
    # Балтика
    ("Финский залив",          "Санкт-Петербург",           1.05),
    ("Выборгский залив",       "Выборг",                    0.95),
    ("Невская губа",           "Кронштадт",                 1.00),
    ("Калининградский залив",  "Балтийск",                  0.95),
    ("Куршская коса",          "Пионерский",                0.90),
    # Ладога и Онега
    ("Ладога",                 "Сортавала",                 1.00),
    ("Ладожские шхеры",        "Приозерск",                 0.95),
    ("Онежское озеро",         "Петрозаводск",              0.95),
    # Русский Север
    ("Белое море",             "Чупа",                      1.10),
    ("Кандалакшский залив",    "Кандалакша",                1.10),
    ("Онежская губа",          "Соловецкий",                1.15),
    ("Двинская губа",          "Архангельск",               1.05),
    ("Кольский залив",         "Мурманск",                  1.20),
    ("Баренцево море",         "Териберка",                 1.30),
    ("Печорское море",         "Нарьян-Мар",                1.35),
    ("Карское море",           "Диксон",                    1.45),
    # Дальний Восток
    ("Анадырский залив",       "Анадырь",                   1.50),
    ("Берингово море",         "Провидения",                1.55),
    ("Командорские острова",   "Никольское",                1.60),
    ("Камчатка",               "Фриза",                     1.40),
    ("Авачинская губа",        "Петропавловск-Камчатский",  1.35),
    ("Охотское море",          "Магадан",                   1.30),
    ("Шантарские острова",     "Чумикан",                   1.45),
    ("Татарский пролив",       "Ванино",                    1.15),
    ("Сахалин",                "Корсаков",                  1.25),
    ("Курильские острова",     "Южно-Курильск",             1.40),
    ("Залив Терпения",         "Поронайск",                 1.20),
    # Приморье
    ("Залив Петра Великого",   "Славянка",                  1.10),
    ("Владивосток",            "Центральный",               1.15),
    ("Залив Находка",          "Находка",                   1.05),
    # Азовское море
    ("Азовское море",          "Ейск",                      0.80),
    ("Таганрогский залив",     "Таганрог",                  0.80),
    ("Темрюкский залив",       "Темрюк",                    0.85),
    # Чёрное море
    ("Анапская бухта",         "Анапа",                     0.95),
    ("Цемесская бухта",        "Новороссийск",              1.00),
    ("Геленджикская бухта",    "Геленджик",                 1.05),
    ("Большой Сочи",           "Сочи",                      1.20),
    ("Туапсинский залив",      "Туапсе",                    0.95),
    # Каспий
    ("Северный Каспий",        "Астрахань",                 0.85),
    ("Дагестанское побережье", "Махачкала",                 0.90),
]

# --- Модели лодок и их длина ------------------------------------------------
# Длина — свойство модели, а не случайная величина: у Bavaria 45 не бывает
# 13 метров. Там, где модель есть в исходной таблице, берётся длина оттуда
# («Полутонник» 10 м, «Конрад 25RT» 9 м, McGregor 65 — 22 м, Судоход 115 — 28 м),
# для остальных — паспортная LOA, округлённая до метра.
SMALL = [  # дневные выходы, кают мало, гальюна может не быть
    ("Конрад 25RT", 9), ("Конрад 24", 8), ("Полутонник", 10),
    ("Четвертьтонник", 8), ("Картер 30", 9), ("Нева", 9), ("Опал 3", 8),
    ("Ассоль", 8), ("Рикошет 747", 8), ("Дракон", 9),
    ("Beneteau First 31.7", 10), ("Bavaria 32", 10), ("Dehler 29", 9),
    ("Гидра 26", 8), ("Botnia Targa 32", 10), ("Л-6", 11),
]
MID = [  # рабочая лошадка круизного чартера
    ("Bavaria 45", 15), ("Bavaria Cruiser 46", 14), ("Hanse 445", 13),
    ("Hanse 415", 12), ("Beneteau Oceanis 45", 14), ("Beneteau Oceanis 41", 12),
    ("Jeanneau Sun Odyssey 449", 14), ("Jeanneau Sun Odyssey 410", 12),
    ("Dufour 460 Grand Large", 14), ("Elan Impression 40", 12),
    ("Salona 44", 13), ("Alubat Ovni 435", 13), ("Nimbus 405 Coupe", 13),
    ("Botnia Targa 37", 12), ("Galeon 425 HTS", 13), ("Lagoon 42", 13),
    ("Nauticat 515", 16), ("Судоход 88", 17),
]
LARGE = [  # экспедиции и представительские выходы
    ("McGregor 65", 22), ("Судоход 115", 28), ("Swan 65", 20),
    ("Lagoon 620", 19), ("Princess 62", 19), ("Azimut 78", 24),
    ("Ferretti 720", 22), ("Beneteau Oceanis Yacht 62", 19), ("Amel 64", 20),
    ("Таганрог 70", 21), ("Sunseeker 76", 23), ("Jongert 24T", 24),
    ("Hallberg-Rassy 64", 20),
]

NAMES = [
    # Русские
    "Северный ветер", "Полярная звезда", "Медуза", "Чайка", "Альбатрос",
    "Буревестник", "Кассиопея", "Вега", "Арктур", "Сириус", "Аврора",
    "Стрелец", "Ладья", "Поморка", "Странник", "Пилигрим", "Норд-Ост",
    "Зюйд", "Вест", "Грот", "Стаксель", "Спинакер", "Бриз", "Шквал",
    "Штиль", "Прибой", "Волна", "Янтарь", "Гранит", "Валун", "Шхера",
    "Салма", "Луда", "Корга", "Баклан", "Гагара", "Крачка", "Поморник",
    "Кайра", "Нерпа", "Белуха", "Косатка", "Сивуч", "Калан", "Треска",
    "Навага", "Корюшка", "Зубатка", "Палтус", "Сайра", "Мойва", "Рябина",
    "Морошка", "Брусника", "Вороника", "Ягель", "Тундра", "Сопка",
    "Вулкан", "Гейзер", "Фумарола", "Кальдера", "Пурга", "Метель",
    "Позёмка", "Наст", "Торос", "Полынья", "Припай", "Меридиан",
    "Параллель", "Румб", "Кабельтов", "Секстант", "Компас", "Штурвал",
    "Клотик", "Салинг", "Бушприт", "Кливер", "Галс", "Оверштаг",
    "Фордевинд", "Бейдевинд", "Бакштаг", "Левентик", "Одиссея", "Скиталец",
    "Беломор", "Помор", "Варяг", "Русич", "Ушкуй", "Коч", "Шхуна",
    "Ветер странствий", "Белая ночь-2", "Серебряный век",
    # Английские и латинские
    "Odyssey", "Vagabond", "Nomad", "Tramontana", "Mistral", "Sirocco",
    "Levante", "Bora", "Meltemi", "Zephyr", "Aeolus", "Poseidon", "Triton",
    "Nereid", "Calypso", "Circe", "Siren", "Argo", "Nautilus", "Neptune",
    "Orion", "Perseus", "Andromeda", "Lyra", "Cygnus", "Aquila", "Corvus",
    "Delphinus", "Hydra", "Kraken", "Leviathan", "Barracuda", "Marlin",
    "Tarpon", "Sailfish", "Wanderer", "Drifter", "Serenity", "Solitude",
    "Freedom", "Endeavour", "Resolute", "Intrepid", "Valiant", "Gallant",
    "Dauntless", "Tenacity", "Fortitude", "Quicksilver", "Silver Lining",
    "Sea Wolf", "Northern Light", "Midnight Sun", "Blue Hour", "Salty Dog",
    "Wind Chaser", "Storm Petrel", "White Squall", "Aurora Borealis",
    "Far Horizon", "Cold Water", "True North", "Sea Fever", "Gulf Stream",
]

# Фотографии, к которым есть смысловая привязка к месту. Остальным строкам
# кадры раздаются по кругу — соответствия лодке и подписи нет, это макет.
PHOTO_HINTS = {
    "Ладога": "Ladoga._Lauvatsaari",
    "Ладожские шхеры": "%D0%92%D0%B5%D1%82%D0%B5%D1%80_%D0%9B%D0%B0%D0%B4%D0%BE%D0%B3",
    "Онежское озеро": "Onego-2011",
    "Финский залив": "Sailing_in_Gulf_of_Finland%2C_Russia-1",
    "Невская губа": "Sunset_over_Gulf_of_Finland",
    "Белое море": "%D0%91%D0%B5%D0%BB%D0%BE%D0%B5_%D0%BC%D0%BE%D1%80%D0%B5_76",
    "Двинская губа": "%D1%80%D0%B5%D1%87%D0%BD%D0%BE%D0%B9_%D1%8F%D1%85%D1%82",
    "Азовское море": "%D0%9C%D0%B5%D0%B2%D1%8B_%D0%BD%D0%B0_%D0%90%D0%B7%D0%BE%D0%B2%D0%B5",
    "Таганрогский залив": "Taganrog._Yacht_club",
    "Владивосток": "Admiral_nevelskoi",
    "Залив Петра Великого": "176_",
}


def pick_model(rng):
    """Модель определяет длину. Мелких лодок во флоте больше, крупных мало."""
    roll = rng.random()
    if roll < 0.42:
        return rng.choice(SMALL)
    if roll < 0.84:
        return rng.choice(MID)
    return rng.choice(LARGE)


def build_row(rng, name, region, port, coeff, photo, models=None):
    """
    Одна строка таблицы по формулам, откалиброванным на эталонных лодках.

    `models` задаёт пул, из которого берётся корпус, — им пользуется
    expansion.py, чтобы выбрать моторную лодку. Без него пул выбирается
    как раньше, и поведение генератора не меняется.
    """
    model, length = rng.choice(models) if models else pick_model(rng)

    # Гостей — примерно 0.62 на метр длины. Коэффициент снят с эталонных
    # строк: 9 м / 6 гостей, 15 м / 10, 28 м / 16.
    guests = max(2, int(round(length * 0.62)) + rng.choice([-1, 0, 0, 1]))

    # Кают — около половины числа гостей; иногда салон-трансформер вместо каюты.
    cabins = max(1, min(10, int(round(guests / 2.0)) + rng.choice([-1, 0, 0])))

    # Гальюн: на лодках короче 12 м его часто нет вовсе.
    if length < 12:
        toilet = "Нет" if rng.random() < 0.7 else 1
    else:
        toilet = max(1, min(4, round(cabins / 2.5) + rng.choice([0, 0, 1])))

    # Состояние: чем новее и крупнее лодка, тем выше оценка.
    state = rng.choice([3, 3, 4, 4, 4, 5]) if length < 18 else rng.choice([4, 4, 5, 5])

    # Цена за сутки. База откалибрована по шести эталонным строкам:
    # 2200 + состояние*200 + длина*25 даёт 15–64 тыс. ₽ на их размерах.
    base = 2200 + state * 200 + length * 25
    price = int(round(guests * base * coeff / 500.0) * 500)

    # Повар — только там, где есть камбуз и место для него: от 12 метров.
    cook = "Да" if length >= 12 else "Нет"

    # Помощник обязателен на крупных лодках; на малых бывает у сложных
    # классических корпусов — как на десятиметровой «Белой ночи».
    if length >= 14:
        crew = "Капитан и помощник"
    else:
        crew = "Капитан и помощник" if rng.random() < 0.18 else "Капитан"

    return [name, model, crew, cook, region, port, cabins, "%d м" % length,
            guests, toilet, price, state, photo]


def main():
    rng = random.Random(20260724)  # фиксированное зерно — таблица воспроизводима

    photos = [p["url"] for p in
              json.load(open(os.path.join(HERE, "photos.source.json")))]
    pool = TILDA_PHOTOS + photos

    used_names = set(r[0] for r in SEED_ROWS)
    names = [n for n in NAMES if n not in used_names]
    rng.shuffle(names)
    name_iter = iter(names)

    rows = []
    photo_cursor = [0]

    for region, port, coeff in REGIONS:
        existing = [r for r in SEED_ROWS if r[4] == region]
        rows.extend(existing)

        # Три лодки — минимум; на популярных направлениях ставим четвёртую.
        target = 4 if coeff >= 1.15 or region in (
            "Ладога", "Финский залив", "Большой Сочи", "Владивосток") else 3
        need = max(0, target - len(existing))

        for i in range(need):
            name = next(name_iter)
            hint = PHOTO_HINTS.get(region)
            photo = None
            if hint and i == 0:
                photo = next((u for u in pool if hint in u), None)
            if photo is None:
                photo = pool[photo_cursor[0] % len(pool)]
                photo_cursor[0] += 1
            rows.append(build_row(rng, name, region, port, coeff, photo))

    # Строки без фото в исходнике тоже должны получить кадр.
    for row in rows:
        if not row[12]:
            row[12] = pool[photo_cursor[0] % len(pool)]
            photo_cursor[0] += 1

    write_csv(rows)
    write_xlsx(rows)
    # PHOTO-CREDITS.md пишет find-photos.py: он владеет фотопулом,
    # и форма photos.source.json теперь его.

    print("Регионов: %d" % len(REGIONS))
    print("Лодок: %d" % len(rows))
    print("Уникальных названий: %d" % len(set(r[0] for r in rows)))
    print("Фотографий в обороте: %d" % len(set(r[12] for r in rows)))


def write_csv(rows):
    path = os.path.join(HERE, "yachts.csv")
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(COLUMNS)
        writer.writerows(rows)


def write_xlsx(rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Яхты"

    head_fill = PatternFill("solid", fgColor="1F3B4D")
    head_font = Font(bold=True, color="FFFFFF")

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in rows:
        ws.append(row)

    # Ссылка на фото — кликабельная, как в исходном файле.
    for idx in range(2, len(rows) + 2):
        cell = ws.cell(row=idx, column=13)
        if cell.value:
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"
            cell.value = "фото"

    widths = [22, 24, 20, 16, 24, 26, 8, 9, 17, 10, 18, 12, 10]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:M%d" % (len(rows) + 1)
    wb.save(os.path.join(HERE, "yachts.xlsx"))


if __name__ == "__main__":
    main()
